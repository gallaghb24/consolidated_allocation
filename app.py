import streamlit as st
import pandas as pd
from io import BytesIO
from collections import defaultdict

# Excel styling helpers
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("âŒ `openpyxl` is not installed. Please run `pip install openpyxl`." )
    st.stop()

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE)
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
BOLD_FONT = Font(bold=True)

# Column definitions
KEY_COLS = [
    "Store Number", "Store Name", "Address Line 1", "Address Line 2", "City or Town",
    "County", "Country", "Post Code", "Region / Area", "Location Type", "Trading Format",
]
LABELS = [
    "POS Code", "Kit Name", "Project Description", "Part", "Supplier",
    "Brief Description", "Total (inc Overs)", "Total Allocations", "Overs",
]
LABEL_COL_XL = KEY_COLS.index("Trading Format") + 1  # K column (1â€‘based)
ITEM_START_XL = LABEL_COL_XL + 1                      # L column (1â€‘based)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Data helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def extract_alloc(file):
    """Read one allocation export â†’ (DataFrame, meta dict)."""
    df = pd.read_excel(file, header=6, engine="openpyxl")
    df["Store Number"] = pd.to_numeric(df["Store Number"], errors="coerce").astype("Int64")

    raw = pd.read_excel(file, header=None, engine="openpyxl")
    meta = {}
    for col in range(len(KEY_COLS), raw.shape[1]):
        ref = str(raw.iloc[6, col])
        if ref == "nan":
            continue
        meta[ref] = {
            "brief_description": raw.iloc[1, col],
            "overs": 0 if pd.isna(raw.iloc[4, col]) else raw.iloc[4, col],
        }
    return df, meta


def merge_allocations(dfs):
    if not dfs:
        return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    num_cols = [c for c in combined.columns if c not in KEY_COLS]
    combined[num_cols] = combined[num_cols].apply(pd.to_numeric, errors="coerce")
    agg = {c: ("first" if c in KEY_COLS else "sum") for c in combined.columns if c != "Store Number"}
    return combined.groupby("Store Number", as_index=False).agg(agg).sort_values("Store Number").reset_index(drop=True)


def load_brief(file):
    """Load Consolidated Brief into a lookup dictionary.

    The Supplier column is optional; when absent the remaining
    details are still returned for each Brief Ref. Columns are
    matched by name so order does not matter.
    """
    if file is None:
        return {}

    brief = pd.read_excel(file, header=1, engine="openpyxl")

    required = {"Brief Ref", "POS Code", "Project Description", "Part"}
    missing = required - set(brief.columns)
    if missing:
        st.error("Consolidated Brief missing columns: " + ", ".join(missing))
        return {}

    has_supplier = "Supplier" in brief.columns
    out = {}
    for _, row in brief.dropna(subset=["Brief Ref"]).iterrows():
        ref = str(row["Brief Ref"]).strip()
        info = {
            "pos_code": row["POS Code"],
            "project_description": row["Project Description"],
            "part": row["Part"],
        }
        if has_supplier:
            info["supplier"] = row["Supplier"]
        out[ref] = info

    return out

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Workbook builder â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_workbook(df: pd.DataFrame, meta: dict, event_code: str) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        STARTROW = len(LABELS) + 1  # pandas header â†’ Excel rowÂ 11
        df.to_excel(writer, index=False, sheet_name="Master Allocation", startrow=STARTROW)
        ws = writer.sheets["Master Allocation"]

        # RowÂ 1 â€“ Project Ref & Event Code
        ws.cell(row=1, column=1, value="Project Ref").font = BOLD_FONT
        ws.cell(row=1, column=2, value=event_code).font = BOLD_FONT

        # Column widths & hide Câ€“J
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
            if "C" <= col_letter <= "J":
                ws.column_dimensions[col_letter].hidden = True

        # Header rows (2â€‘10)
        item_cols = [c for c in df.columns if c not in KEY_COLS]
        for r_off, label in enumerate(LABELS):
            row_num = 2 + r_off
            lh = ws.cell(row=row_num, column=LABEL_COL_XL, value=label)
            lh.alignment = Alignment(wrap_text=(row_num in (5, 7)), vertical="center")
            lh.fill = ORANGE_FILL
            lh.font = BOLD_FONT
            lh.border = THIN_BORDER
            for idx, item in enumerate(item_cols):
                cell = ws.cell(row=row_num, column=ITEM_START_XL + idx)
                data = meta.get(item, {})
                overs = data.get("overs", 0)
                total = df[item].fillna(0).sum()
                if label == "POS Code":
                    cell.value = data.get("pos_code", "")
                elif label == "Project Description":
                    cell.value = data.get("project_description", "")
                elif label == "Part":
                    cell.value = data.get("part", "")
                elif label == "Supplier":
                    cell.value = data.get("supplier", "")
                elif label == "Brief Description":
                    cell.value = data.get("brief_description", "")
                elif label == "Total (inc Overs)":
                    cell.value = total + overs
                elif label == "Total Allocations":
                    cell.value = total
                elif label == "Overs":
                    cell.value = overs
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(row_num in (5, 7)))
                cell.border = THIN_BORDER

        # Style pandas header (Excel rowÂ 11)
        header_excel_row = STARTROW + 1
        for col_idx in range(1, ws.max_column + 1):
            hcell = ws.cell(row=header_excel_row, column=col_idx)
            hcell.fill = ORANGE_FILL
            hcell.font = BOLD_FONT
            hcell.border = THIN_BORDER

        # Data rows
        for row in ws.iter_rows(min_row=header_excel_row + 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                if c.column >= ITEM_START_XL:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = THIN_BORDER

    buffer.seek(0)
    return buffer

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ Streamlit UI â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

st.set_page_config(page_title="Superdrug Consolidated Allocation Builder", layout="wide")

st.title("Superdrug Consolidated Allocation Builder")

st.markdown("""**StepÂ 1 â€“ Upload all allocation exports together** â€“ [download them here](https://superdrug.aswmediacentre.com/ArtworkPrint/ArtworkPrintReport/ArtworkPrintReport?reportId=1149)  
**StepÂ 2 â€“ Upload the Consolidated Brief complete with Supplier for each line (optional)**  
**StepÂ 3 â€“ Enter the Event Code (required)**  
**StepÂ 4 â€“ Download the Consolidated Allocation**""")

alloc_files = st.file_uploader("Allocation exports (.xlsx)", type=["xlsx"], accept_multiple_files=True)
brief_file = st.file_uploader("Consolidated Brief (.xlsx)", type=["xlsx"], key="brief")
event_code = st.text_input("Event Code (e.g. E0625)")

if not alloc_files:
    st.info("Please upload at least one allocation export.")
    st.stop()
if not event_code.strip():
    st.warning("Event Code is required.")
    st.stop()

# Merge process
progress = st.progress(0)
all_dfs, meta = [], defaultdict(dict)
for idx, up in enumerate(alloc_files, start=1):
    df_part, meta_part = extract_alloc(up)
    all_dfs.append(df_part)
    for k, v in meta_part.items():
        meta.setdefault(k, {}).update(v)
    progress.progress(idx / len(alloc_files))
progress.empty()

for ref, info in load_brief(brief_file).items():
    meta.setdefault(ref, {}).update(info)

master_df = merge_allocations(all_dfs)

workbook_bytes = build_workbook(master_df, meta, event_code.strip())

# Success message and download
lines_count = master_df.shape[1] - len(KEY_COLS)
st.success(f"Consolidated {lines_count} lines Ã— {master_df.shape[0]} stores.")

st.dataframe(master_df.head(50), use_container_width=True)

st.download_button(
    label="ðŸ“¥ Download the Consolidated Allocation",
    data=workbook_bytes,
    file_name=f"{event_code.strip()}_Consolidated_Allocation.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
